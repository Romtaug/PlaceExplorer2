receiver_email = ["@gmail.com"]

##########################################################################

# Si la clé API Google Places a expiré ou n'est plus valide :
# 1. Accédez à Google Cloud Platform : https://console.cloud.google.com/
# 2. Créez un nouveau projet ou sélectionnez un projet existant.
# 3. Dans le menu de navigation, allez dans "API et services" > "Bibliothèque".
# 4. Recherchez et sélectionnez l'API "Places API" (Google Places API).
# 5. Cliquez sur "Activer" pour activer l'API pour votre projet.
# 6. Une fois l'API activée, allez dans "API et services" > "Identifiants" pour créer une clé API.
# 7. Copiez la clé API générée et collez-la dans la variable `api_key` du programme ci-après :
api_key = '' # API Events
# Si la console vous empêche de créer un nouveau projet (quota atteint ou "plein") :
# ➜ Supprimez un projet inutile pour faire de la place :
#    1. Accédez à Google Cloud Platform : https://console.cloud.google.com/
#    2. Cliquez sur le menu ☰ (en haut à gauche).
#    3. Allez dans "IAM et administration" > "Gérer les ressources".
#    4. Sélectionnez le projet à supprimer.
#    5. Cliquez sur ⋮ à droite puis sur "Supprimer".
#    6. Tapez l'ID du projet pour confirmer.
#    7. Le projet sera désactivé immédiatement, puis supprimé définitivement après 30 jours.

# Pour autoriser l'accès à l'email via des applications moins sécurisées :
# 1. Connectez-vous à votre compte Gmail : https://mail.google.com/
# 2. Cliquez sur l'icône de votre profil en haut à droite, puis sur "Gérer votre compte Google".
# 3. Dans le menu, allez dans "Sécurité".
# 4. Sous la section "Accès de l'application moins sécurisée", actAivez l'option "Accès autorisé".
#    (Note : Cette option peut ne pas être disponible si l'authentification à deux facteurs est activée.
#     Dans ce cas, vous devrez créer un mot de passe spécifique à l'application.)
# 5. Utilisez cette adresse email et le mot de passe associé dans le programme.

import re
import sys

# Fonction de validation des emails
def is_valid_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

# Validation des emails
for email in receiver_email:
    if not is_valid_email(email):
        print(f"❌ Adresse email invalide : {email}")
        sys.exit(1)

print("Toutes les adresses email sont valides.")


###########################################################################

import importlib
import subprocess
import sys

# Modules de la librairie standard qu'on n'a pas besoin d'installer
standard_libs = {'time', 'smtplib', 'email', 'os', 'unicodedata'}

# Modules à vérifier
modules = [
    'requests', 'pandas', 'openpyxl'
]

# Vérification et installation si besoin
for module in modules:
    if module not in standard_libs:
        try:
            importlib.import_module(module)
        except ImportError:
            print(f"{module} non installé, installation en cours...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", module])

# Importations
import requests
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import pandas as pd
import openpyxl
import shutil
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import Font

print("Tous les modules sont importés avec succès.")


def apply_hyperlink_styles(file_path):
    """
    Applique le style bleu et souligné pour les colonnes contenant des hyperliens
    tout en affichant le lien complet dans les cellules.
    """
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):  # Ignorer les en-têtes
            cell = row[5]  # Supposons que la colonne "Maps" est la 6e colonne
            if cell.value and "https://" in cell.value:
                cell.hyperlink = cell.value  # Convertir en hyperlien cliquable
                cell.font = Font(color="0000FF", underline="single")  # Appliquer le style
                # Afficher le lien complet dans la cellule
                cell.value = cell.value
    wb.save(file_path)
    print(f"✅ Hyperlink styles applied with full links to: {file_path}")

def remove_invalid_rows(file_path, location):
    """
    Supprime les lignes dans toutes les feuilles Excel où les adresses
    ne se terminent pas par le pays spécifié.
    """
    # Extraire le pays uniquement (dernier mot après une éventuelle virgule)
    country = location.split(",")[-1].strip()  # On garde uniquement le pays
    print(f"Filtrage des lignes se terminant par : {country}")

    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_to_delete = []

        # Identifier les lignes à supprimer
        for row in ws.iter_rows(min_row=2):  # Ignorer la première ligne (en-têtes)
            cell = row[1]  # Supposons que les adresses sont dans la 2ème colonne
            if cell.value and isinstance(cell.value, str):
                # Normaliser la valeur de la cellule
                cell_value_normalized = " ".join(cell.value.split())
                # Supprimer si l'adresse ne se termine pas par le pays
                if not cell_value_normalized.endswith(country):
                    rows_to_delete.append(cell.row)

        # Supprimer les lignes identifiées
        for row_idx in sorted(set(rows_to_delete), reverse=True):
            ws.delete_rows(row_idx)

    wb.save(file_path)
    print(f"✅ Lignes se terminant par '{country}' conservées dans le fichier : {file_path}")

def normalize_location(location):
    """
    Normalise une entrée de localisation en la forçant au format "Ville, Pays" ou "Pays".
    Redemande l'entrée si le format est incorrect.
    """
    while True:
        # Supprime les espaces multiples et normalise
        location = " ".join(location.strip().split())
        
        # Vérifie si l'entrée correspond à "Pays" ou "Ville, Pays"
        if "," in location:
            parts = location.split(",")
            if len(parts) == 2:
                city, country = parts[0].strip(), parts[1].strip()
                if city and country:
                    # Convertir les deux parties en forme correcte
                    normalized = f"{city.title()}, {country.title()}"
                    return normalized
        elif location.isalpha():
            # Si seulement un mot (ex. "France")
            return location.title()

        # Si le format est incorrect, redemander à l'utilisateur
        print("❌ Format incorrect. Veuillez entrer une localisation au format 'Ville, Pays' ou 'Pays'.")
        location = input("Please enter the location in english (ex: Paris, France; or : France) : ").strip()

def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(file_path)

# Fonction pour normaliser les noms de fichiers (supprimer les accents)
def normalize_filename(filename):
    return ''.join(
        c for c in unicodedata.normalize('NFD', filename)
        if unicodedata.category(c) != 'Mn'
    )

def get_place_details(place_id, api_key):
    details_url = "https://maps.googleapis.com/maps/api/place/details/json"
    details_params = {
        'place_id': place_id,
        'fields': 'name,formatted_address,rating,user_ratings_total,price_level,international_phone_number,website',
        'key': api_key
    }
    response = requests.get(details_url, params=details_params)
    if response.status_code == 200:
        return response.json().get('result', {})
    else:
        print(f"Failed to fetch details for place_id: {place_id}")
        return None

# Fonction pour récupérer les lieux via Google Places API
import requests
import time

def extract_city_from_address(full_address):
    """
    Extrait la ville d'une adresse en prenant :
    - L'avant-dernier élément si des virgules existent.
    - Le premier mot si aucune virgule.
    - Gère les cas de 'Unnamed Road'.
    """
    address_parts = full_address.split(", ")

    if "Unnamed Road" in full_address:
        # Si l'adresse contient "Unnamed Road", on prend la partie après
        city = address_parts[-2] if len(address_parts) > 2 else address_parts[-1]
    elif len(address_parts) > 2:
        # Si on a au moins 2 virgules, on prend l'avant-dernier élément
        city = address_parts[-2]
    elif len(address_parts) == 2:
        # Si une seule virgule, prendre le premier élément (souvent la ville)
        city = address_parts[0]
    else:
        # Si pas de virgule, prendre le premier mot
        city = full_address.split(" ")[0]

    return city

def search_places(api_key, location, category):
    """
    Recherche les 20 lieux les plus populaires dans une localisation et une catégorie donnée via l'API Google Places.
    Ajoute une colonne 'City' en extrayant correctement la ville de l'adresse.
    """
    import requests
    import time

    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {'query': f'{category} in {location}', 'key': api_key}
    places = []

    response = requests.get(url, params=params)
    if response.status_code == 200:
        results = response.json()
        places = results.get('results', [])[:20]  # ✅ On prend juste les 20 premiers lieux (page 1)
    else:
        print(f"Erreur {response.status_code} lors de la récupération des données pour {category}")
        return []

    detailed_places = []
    for place in places:
        place_id = place.get('place_id')
        if not place_id:
            continue

        details = get_place_details(place_id, api_key)
        if details:
            full_address = details.get('formatted_address', 'Not specified')
            city = extract_city_from_address(full_address)

            detailed_places.append({
                'City': city,
                'Address': full_address,
                'Name': details.get('name', 'Not specified'),
                'Total Reviews': details.get('user_ratings_total', 0),
                'Rating (on 5)': details.get('rating', 'Not rated'),
                'Price Level': {
                    0: "Free",
                    1: "+",
                    2: "++",
                    3: "+++",
                    4: "++++"
                }.get(details.get('price_level', None), 'Not specified'),
                'Maps': f'=HYPERLINK("https://www.google.com/maps/place/?q=place_id:{place_id}", '
                        f'"https://www.google.com/maps/place/?q=place_id:{place_id}")',
                'Phone': details.get('international_phone_number', 'Not available')
            })

    return detailed_places

# Fonction pour ajuster la largeur des colonnes
def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Utiliser openpyxl.utils.get_column_letter si besoin
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Ajout d'une marge
            ws.column_dimensions[col_letter].width = adjusted_width
    wb.save(file_path)

# Modifier la fonction create_excel_file
def create_excel_file(api_key):
    print("Please enter the location in english (ex : Paris, France; or : France) :")
    location = input().strip()
    normalized_location = normalize_location(location)
    location = " ".join(location.split())

    print("Entrez votre mois de vacances (ex : Août, 2025):")
    vacation_month = input().strip()
    vacation_month = " ".join(vacation_month.split())
    vacation_month = vacation_month.replace(",", "-").replace(" ", "")

    # Remplacer les virgules par des tirets et supprimer les espaces dans le nom du fichier
    sanitized_location = location.replace(",", "-").replace(" ", "")

    # Générer le chemin temporaire
    script_dir = os.path.dirname(os.path.abspath(__file__))  # dossier du script
    file_name = f"{sanitized_location}_{vacation_month}.xlsx"
    file_path = os.path.join(script_dir, file_name)

    writer = pd.ExcelWriter(file_path, engine='openpyxl')

    categories = {
        # Découvertes
        'historical_sites': '🏰 Sites historiques',
        'museums': '🖼️ Musées',
        'churches': '⛪ Églises',
        'cultural_centers': '🎭 Centres culturels',
        'hiking_trails': '🥾 Sentiers de randonnée',
        # Restauration
        'restaurants': '🍴 Restaurants',
        'bars': '🍹 Bars',
        # Détente
        'parks': '🌳 Parcs',
        'beaches': '🏖️ Plages',
        'lakes': '🏞️ Lacs',
        # Divertissements
        'concert_halls': '🎶 Salles de concert',
        'nightclubs': '💃 Boîtes de nuit',
        'movie_theaters': '🎬 Cinémas',
        'stadiums': '🏟️ Stades',
        # Shopping
        'markets': '🌽 Marchés',
        'boutiques': '🛍️ Boutiques',
        'supermarkets': '🛒 Supermarchés',
        # Activités
        'festivals' : '🎉 Festivals',
        'amusement_parks': '🎢 Parcs d\'attractions',
        'zoos': '🐘 Zoos',
        'aquariums': '🐠 Aquariums',
        'mountain_resorts': '🏔️ Stations de montagne',
        # Sport
        'bike_rentals' : '🚴 Locations de vélos',
        'campgrounds': '🏕️ Campings',
        'sports_centers': '🏋️‍♂️ Centres sportifs',
        'spas': '💆‍♀️ Spas',
        'gym': '🏋️‍♀️ Salles de sport',
        # Transports
        'train_stations': '🚆 Gares',
        'airports': '✈️ Aéroports',
        # Éducation
        'schools': '🏫 Écoles',
        # Santé
        'hospitals': '🏥 Hôpitaux',
    }

    for category, description in categories.items():
        print(f"Fetching data for category: {category}")
        data = search_places(api_key, location, category)
        if data:
            df = pd.DataFrame(data)

            # Trier les données par nombre de commentaires
            df = df.sort_values(by='Total Reviews', ascending=False)

            # Ajuster les noms des feuilles
            sheet_name = description if len(description) <= 31 else description[:31]  # Excel limite les noms de feuilles à 31 caractères

            # Écrire dans Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()  # Fermer l'écrivain pour sauvegarder le fichier
    
    #apply_hyperlink_styles(file_path)
    
    adjust_column_width(file_path)  # Ajuster la largeur des colonnes
    print(f"Excel file created with clickable links: {file_path}")
    return file_path, location, vacation_month

# Fonction pour envoyer l'e-mail avec fichier joint et images
def send_email_with_excel(sender_email, password, receiver_emails, subject, body, file_path, image_paths):
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)  # Ajouter tous les destinataires
    msg['Subject'] = subject

    # Ajouter le texte du corps de l'e-mail
    msg.attach(MIMEText(body, 'plain'))

    # Ajouter le fichier Excel en pièce jointe
    try:
        with open(file_path, 'rb') as file:
            part = MIMEBase('application', "octet-stream")
            part.set_payload(file.read())
        encoders.encode_base64(part)
        normalized_name = normalize_filename(os.path.basename(file_path))
        part.add_header('Content-Disposition', f'attachment; filename="{normalized_name}"')
        msg.attach(part)
    except FileNotFoundError:
        print(f"Fichier non trouvé : {file_path}")
        return

    # Ajouter les images en pièce jointe
    for image_path in image_paths:
        if os.path.exists(image_path):
            with open(image_path, 'rb') as img:
                img_part = MIMEBase('application', "octet-stream")
                img_part.set_payload(img.read())
            encoders.encode_base64(img_part)
            img_part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(image_path)}"')
            msg.attach(img_part)
        else:
            print(f"Image non trouvée : {image_path}")

    # Envoi de l'e-mail
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_emails, msg.as_string())  # Envoi à plusieurs destinataires
            print(f"E-mail envoyé avec succès à {', '.join(receiver_emails)}")
    except Exception as e:
        print(f"Erreur lors de l'envoi de l'e-mail : {e}")

# Script principal
if __name__ == "__main__":
    # Clés API et configuration
    api_key = api_key
    sender_email = "@cy-tech.fr"
    password = ""
    receiver_email = receiver_email

    # Création du fichier Excel
    try:
        excel_file, location, vacation_month = create_excel_file(api_key)
        remove_invalid_rows(excel_file, location)
        adjust_column_width(excel_file)
    except Exception as e:
        print(f"Erreur lors de la création du fichier Excel : {e}")
        exit()

    # Nettoyage des variables pour éviter les problèmes dans le sujet ou les noms de fichiers
    location_cleaned = location.replace(",", "-").replace(" ", "")
    vacation_month_cleaned = vacation_month.replace(" ", "-")

    # Vérifier si le fichier Excel existe
    if not os.path.exists(excel_file):
        print(f"Le fichier Excel n'a pas été trouvé : {excel_file}")
        exit()

    # Images à joindre
    import os

    # Obtenir le chemin absolu du script courant
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Construire les chemins absolus des images
    image_paths = [
        os.path.join(script_dir, "Image", "logo.png"),
        os.path.join(script_dir, "Image", "travel.png"),
        os.path.join(script_dir, "Image", "qrcode.png")
    ]

    # Texte de l'e-mail
    email_body = f"""Bienvenue à bord de Place Explorer !

Découvrez {location} ! C'est une destination rêvée pour des aventures inoubliables. Votre guide inclut :
    - Les lieux incontournables à visiter
    - Un plan d'organisation pour votre voyage

🗺 Étapes pour organiser votre voyage :
    1. Ouvrez le fichier Excel attaché avec Google Sheet en cliquant une fois dessus
    2. Consultez chaque feuille pour explorer les meilleurs options par catégorie
    3. Planifiez vos activités (par exemple 2/jours par feuille) sur : https://www.google.com/mymaps un calque par ville
    4. Si le réseau est payant à l’étranger, utilisez Google Maps hors connexion : https://support.google.com/maps/answer/6291838?hl=fr
    5. Envoyez-vous votre parcours du jour (en fonction de la proximité) sur WhatsApp ou via Google Docs pour le garder à portée de main sur votre téléphone : https://web.whatsapp.com, https://wa.me, ou https://docs.google.com

🌍 Liens utiles :
    - ✈ Pour trouver les vols les moins chers et obtenir des indemnisations en cas de retard : https://www.skyscanner.fr/ ou https://www.airhelp.com/fr/
    - 🚅 Pour comparer tous les moyens de transport : https://www.rome2rio.com/
    - 🏠 Pour réserver votre hébergement et véhicule : https://www.airbnb.com et https://www.booking.com
    - 🖍 Pour des avis et recommandations : https://www.tripadvisor.com
    - 🗺 Pour réserver des activités locales : https://www.getyourguide.fr/
    - 📞 Pour trouver des eSIM à moindre coût à l'étranger : https://www.airalo.com/fr
    - 💳 Pour dépenser sans aucuns frais de change et gagner 200 € à l'ouverture : https://revolut.com/referral/?referral-code=romainavh3!DEC1-24-VR-FR

🤖 Copiez ce prompt sur https://chatgpt.com pour enrichir votre expérience :

"Je cherche des expériences et activités extraordinaires à {location_cleaned} en {vacation_month_cleaned}, fais un top 20 des incontournables durant cette période (événements, activités, monuments, restaurants, quartiers) et un top 10 des villes à visiter autour avec le temps de trajet, indique les démarches administratives nécessaires (documents, visas, vaccins), les précautions à prendre (arnaques, numéros d'urgence), les coûts approximatifs, la météo moyenne, les événements locaux, et des astuces pour se déplacer, respecter les coutumes, et profiter au maximum."
        
Nous espérons que vous passerez un moment incroyable. Bon voyage ! ✈
N'hésitez pas à faire un don via PayPal à l'adresse romtaug@gmail.com si cela vous a aidé.
Accédez à notre outil pour travailler à l'étranger : https://bordeuroconnect.netlify.app/"""

# Objet de l'email formaté
subject = f"🌍 PlaceExplorer : Les Meilleurs Lieux Destination {location_cleaned} en {vacation_month_cleaned}"

try:
    # Envoi de l'e-mail avec fichier Excel et images
    send_email_with_excel(sender_email, password, receiver_email, subject, email_body, excel_file, image_paths)

    # Déplacer le fichier dans le dossier "Content" situé à côté du script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    content_dir = os.path.join(script_dir, "Content")

    if not os.path.exists(content_dir):
        os.makedirs(content_dir)
        
    destination = os.path.join(content_dir, os.path.basename(excel_file))
    shutil.move(excel_file, destination)
    print(f"✅ Fichier déplacé dans le dossier Content : {destination}")

except Exception as e:
    print(f"❌ Erreur lors de l'envoi de l'e-mail ou du déplacement du fichier : {e}")
